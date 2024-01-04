from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def check_pinterest_account(email, result_dict):
    url = 'https://in.pinterest.com/'
    driver = webdriver.Chrome()

    try:
        driver.get(url)
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="mweb-unauth-container"]/div/div/div[1]/div/div[2]/div[2]/button/div/div'))
        )
        login_button.click()

        email_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'email'))
        )

        email_input.send_keys(email)
        email_input.send_keys(Keys.RETURN)

        time.sleep(2)

        error_message = None
        try:
            error_message = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="email-error"]/div/div/div[2]'))
            )
        except:
            pass

        if error_message:
            result_dict[email]["Pinterest"] = "❌"
        else:
            result_dict[email]["Pinterest"] = "✔"

    except Exception as e:
        print(f"EXCEPTION: {str(e)}")
        result_dict[email]["Pinterest"] = "Error"

    finally:
        driver.quit()

def check_spotify_account(email, result_dict):
    url = 'https://www.spotify.com/in-en/signup'
    driver = webdriver.Chrome()

    try:
        driver.get(url)

        email_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="username"]'))
        )

        email_input.send_keys(email)
        email_input.submit()

        try:
            error_message = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.XPATH, '//div[@class="Wrapper-sc-62m9tu-0 POdTa encore-warning-set AlreadyInUseBanner__StyledBanner-sc-1j4rkgm-0 jMBpIH"]'))
            )
            result_dict[email]["Spotify"] = "❌"
        
        except TimeoutException:
            next_button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, '//button[@data-testid="submit"]'))
            )
            next_button.click()
            result_dict[email]["Spotify"] = "✔"

    except Exception as e:
        print(f"EXCEPTION: {str(e)}")
        result_dict[email]["Spotify"] = "Error"
    
    finally:
        driver.quit()

def check_quora_account(email, result_dict):
    url = 'https://www.quora.com/'
    driver = webdriver.Chrome()

    try:
        driver.get(url)

        email_input = driver.find_element(By.ID, 'email')
        email_input.send_keys(email)
        email_input.send_keys(Keys.RETURN)

        time.sleep(2)

        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')

        error_message_div = soup.find('div', {'class': 'qu-color--red_error'})
        if error_message_div:
            result_dict[email]["Quora"] = "❌"
        else:
            result_dict[email]["Quora"] = "✔"

    except Exception as e:
        print(f"EXCEPTION: {str(e)}")
        result_dict[email]["Quora"] = "Error"

    finally:
        driver.quit()

excel_file = 'emails.xlsx'
df = pd.read_excel(excel_file, header=None)

results = {}

for index, row in df.iterrows():
    print(f"\nChecking row {index + 1}:")

    # Iterate through the emails in the current row
    for email in row:
        print(f" Checking email: {email}")
        if email not in results:
            results[email] = {"Pinterest": "", "Spotify": "", "Quora": ""}
        check_quora_account(email, results)
        check_pinterest_account(email, results)
        check_spotify_account(email, results)

# Create a new workbook
wb = Workbook()
ws = wb.active

# Header row
header_row = ["Email", "Pinterest", "Spotify", "Quora"]
ws.append(header_row)

# Populate the data and styles
for email, result in results.items():
    row_data = [email] + list(result.values())
    ws.append(row_data)

# Apply styling
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = Font(bold=True)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value == "✔":
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            cell.font = Font(color="FFFFFF")  # White text for green cells
        elif cell.value == "❌":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF")  # White text for red cells

# Save the results to an output file
output_file = 'output.xlsx'
wb.save(output_file)
